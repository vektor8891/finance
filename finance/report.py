#!/usr/bin/env python
import os

import pandas as pd

import finance.dataframe as d
import finance.validate as v
import typing as t


def summarize_transactions(df: pd.DataFrame):
    d.has_column(df, "Date", raise_error=True)
    df["Month"] = pd.DatetimeIndex(df["Date"]).month
    df.sort_values(by=["Date"], inplace=True, ignore_index=True)
    return df


def summarize_categories(df: pd.DataFrame):
    cat_columns = ["CategoryType", "CategoryName"]
    d.has_columns(df, cat_columns + ["Month", "AmountUSD"], raise_error=True)
    df_cat = df.pivot_table(
        index=cat_columns,
        columns=["Month"],
        values=["AmountUSD"],
        aggfunc="sum",
    )
    df_cat.columns = df_cat.columns.droplevel()
    df_cat.columns.name = None
    df_cat = df_cat.fillna(0).reset_index()
    return df_cat


def add_usd_amount(year: int, df: pd.DataFrame):
    d.has_column(df, "Amount", raise_error=True)
    d.has_column(df, "Currency", raise_error=True)

    fx_rates = v.get_fx_rates(year)

    def convert_to_usd(row: pd.Series):
        amount = row["Amount"]
        ccy = row["Currency"]
        return v.convert_amount(amount, ccy, "USD", fx_rates)

    df.loc[:, "AmountUSD"] = df.apply(lambda x: convert_to_usd(x), axis=1)
    return df


def summarize_months(year: int, df: pd.DataFrame):
    df = add_usd_amount(year, df)
    d.has_column(df, "Date", raise_error=True)
    df["Month"] = pd.DatetimeIndex(df["Date"]).month
    df.sort_values(by=["Date"], inplace=True, ignore_index=True)
    df_sum = summarize_categories(df)
    return df, df_sum


def get_balance(year: int):
    import finance.functions as f

    df = d.parse_csv(year, "settings", "accounts.csv")
    df.rename(columns={"InitialBalance": "Amount"}, inplace=True)
    df = add_usd_amount(year, df)
    df["AmountUSD"] = df["AmountUSD"].round()
    del df['Amount']
    df_balance = df.groupby(["AccountType", "AccountCategory"]).sum()
    f_path = f.get_path(year, "output", "balance.xlsx")
    df_balance.reset_index().to_excel(f_path, index=False)


def get_pnl(year: int, df: pd.DataFrame):
    import finance.functions as f

    df["LivingExpense"] = ~df["CategoryName"].str.startswith("Savings") & \
                          ~df["CategoryName"].str.startswith("Donation") & \
                          ~df["CategoryName"].str.startswith("Sunk Costs")
    df_avg = df.pivot_table(
        index=["CategoryType", "LivingExpense", "CategoryName"],
        values=["AmountUSD"],
        aggfunc="sum",
    )
    months = len(df["Month"].unique())
    df_avg["AmountUSD"] = abs(round(df_avg["AmountUSD"] / months))
    f_path = f.get_path(year, "output", "pnl.xlsx")
    df_avg.to_excel(f_path)


def save_results(df: pd.DataFrame, df_sum: pd.DataFrame, f_path: str):
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.load_workbook(f_path)

    def save_worksheet(workbook: openpyxl.workbook.workbook.Workbook, sheet_ind: int, sheet_name: str,
                       data: pd.DataFrame, file_path: str):
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        worksheet = workbook.create_sheet(sheet_name, sheet_ind)
        for _, ws_row in enumerate(dataframe_to_rows(data, header=True, index=False)):
            worksheet.append(ws_row)
        workbook.save(file_path)

    # Save transactions
    df_sorted = df[
        ["Month", "CategoryName", "Comment", "Date", "Account", "Amount", "Currency", "Details"]].sort_values(
        by=["Month", "CategoryName", "Date"], ascending=[False, True, True])
    save_worksheet(workbook=wb, sheet_name="Transactions", sheet_ind=1, data=df_sorted, file_path=f_path)

    # Save summary
    save_worksheet(workbook=wb, sheet_name="Summary", sheet_ind=1, data=df_sum, file_path=f_path)
